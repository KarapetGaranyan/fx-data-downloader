from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import pandas as pd
import datetime
import numpy as np
import os
import json
import MetaTrader5 as mt
import xlrd, requests
import zipfile

def get_close_prices():
    mt.initialize()
    login = 9999778737
    password = 'Zv*8MiCq'
    server = 'MetaQuotes-Demo'
    mt.login(login, password, server)
    symbols = ["EURUSD", "GBPUSD", "AUDUSD", "USDCAD", "USDJPY", "XAUUSD", "XAGUSD"]
    closed_price = [mt.copy_rates_from_pos(i, mt.TIMEFRAME_D1, 1, 1)[0][1] for i in symbols]
    mt.shutdown()
    closed_price[0], closed_price[1], closed_price[2], closed_price[3], closed_price[4], closed_price[5], closed_price[6] = round(closed_price[0] * 10000), round(closed_price[1] * 1000), round(closed_price[2] * 10000), round(1/ (closed_price[3])*10000), round(1/ (closed_price[4]) *1000000), round(closed_price[5]), round(closed_price[6] * 100)
    return closed_price

class Int64Encoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.int64):
            return int(obj)
        return super(Int64Encoder, self).default(obj)

def get_data(currency, option_type, coefficient, price_close):
    call_start_A, call_stop_A, put_start_A, put_stop_A = '', '', '', ''
    wb = load_workbook("new version/" + currency + ".xlsx")
    ws = wb.active
    call_A, call_I, call_J, put_A, put_I, put_J = [],[],[],[],[],[]
    for cell in ws:
        for x in cell:
            if x.value == option_type:
                call_start_A = 'A' + str(x.row + 3)
                call_start_I = 'I' + str(x.row + 3)
                call_start_J = 'J' + str(x.row + 3)
                break
            if call_start_A != "" and call_stop_A == '' and x.value == 'TOTALS':
                call_stop_A = 'A' + str(x.row - 1)
                call_stop_I = 'I' + str(x.row - 1)
                call_stop_J = 'J' + str(x.row - 1)
                break
            if call_start_A != "" and call_stop_A != '' and put_start_A == '' and x.value == 'Strike':
                put_start_A = 'A' + str(x.row + 1)
                put_start_I = 'I' + str(x.row + 1)
                put_start_J = 'J' + str(x.row + 1)
                break
            if call_start_A != "" and call_stop_A != '' and put_start_A != '' and put_stop_A == '' and x.value == 'TOTALS':
                put_stop_A = 'A' + str(x.row - 1)
                put_stop_I = 'I' + str(x.row - 1)
                put_stop_J = 'J' + str(x.row - 1)
                break
    call_a, call_i, call_j, put_a, put_i, put_j = [], [], [], [], [], []
    range_call_A = ws[call_start_A:call_stop_A]
    for cell in range_call_A:
        for x in cell:
            x.value.replace(",", "").replace("'", "")
            call_a.append(x.value)
    range_call_I = ws[call_start_I:call_stop_I]
    for cell in range_call_I:
        for x in cell:
            call_i.append(x.value)
    range_call_J = ws[call_start_J:call_stop_J]
    for cell in range_call_J:
        for x in cell:
            call_j.append(x.value)

    range_put_A = ws[put_start_A:put_stop_A]
    for cell in range_put_A:
        for x in cell:
            put_a.append(x.value)
    range_put_I = ws[put_start_I:put_stop_I]
    for cell in range_put_I:
        for x in cell:
            put_i.append(x.value)
    range_put_J = ws[put_start_J:put_stop_J]
    for cell in range_put_J:
        for x in cell:
            put_j.append(x.value)

    call_a = [int(i.replace(",", "")) for i in call_a]  # price call
    call_i = [int(i.replace(",", "")) for i in call_i]  # strike call
    call_j = [int(i.replace(",", "")) for i in call_j]  # delta call

    put_a = [int(i.replace(",", "")) for i in put_a]  # price put
    put_i = [int(i.replace(",", "")) for i in put_i]  # strike put
    put_j = [int(i.replace(",", "")) for i in put_j]  # delta put

    # STRIKE SIP
    call_df = {
        'Strike': call_a,
        'At Close': call_i,
        'Change': call_j,
    }

    put_df = {
        'Strike': put_a,
        'At Close': put_i,
        'Change': put_j,
    }
    df_call = pd.DataFrame(call_df)
    df_call_less = df_call[df_call['Strike'] < price_close]
    df_put = pd.DataFrame(put_df)
    df_put_more = df_put[df_put['Strike'] > price_close]

    if currency == "EUR" or currency == "GBP" or currency == "AUD" or currency == "XAU" or currency == "XAG":
        up_level = (((df_call['Strike'] * df_call['At Close']).sum() / (df_call['At Close'].sum()))) / coefficient
        down_balance_level = (((df_call_less['Strike'] * df_call_less['At Close']).sum() / (df_call_less['At Close'].sum()))) / coefficient
        down_level = (((df_put['Strike'] * df_put['At Close']).sum() / (df_put['At Close'].sum()))) / coefficient
        up_balance_level = (((df_put_more['Strike'] * df_put_more['At Close']).sum() / (df_put_more['At Close'].sum()))) / coefficient
        red_balance_level = ((((df_call_less['Strike'] * df_call_less['At Close']).sum() + (df_put_more['Strike'] * df_put_more['At Close']).sum()) / (df_call_less['At Close'].sum() + df_put_more['At Close'].sum()))) / coefficient

    if currency == "CAD" or currency == "JPY":
        up_level = (1 / ((df_put['Strike'] * df_put['At Close']).sum() / (df_put['At Close'].sum()))) * coefficient
        down_level =  (1 / ((df_call['Strike'] * df_call['At Close']).sum() / (df_call['At Close'].sum()))) * coefficient
        up_balance_level = (1 / ((df_call_less['Strike'] * df_call_less['At Close']).sum() / (df_call_less['At Close'].sum()))) * coefficient
        down_balance_level = (1 / ((df_put_more['Strike'] * df_put_more['At Close']).sum() / (df_put_more['At Close'].sum()))) * coefficient
        red_balance_level = ( 1 /(((df_call_less['Strike'] * df_call_less['At Close']).sum() + (df_put_more['Strike'] * df_put_more['At Close']).sum()) / (df_call_less['At Close'].sum() + df_put_more['At Close'].sum()))) * coefficient

    # print(up_level, down_level, up_balance_level, down_balance_level, red_balance_level)
    # FOB
    # Опционы в деньгах
    opt_in_money_call_i = df_call_less['At Close'].sum()
    opt_in_money_call_j = df_call_less['Change'].sum()
    opt_in_money_put_i = df_put_more['At Close'].sum()
    opt_in_money_put_j = df_put_more['Change'].sum()

    # Опционы без денег
    df_call_more= df_call [df_call['Strike'] > price_close]
    df_put_less = df_put[df_put['Strike'] < price_close]

    opt_without_money_call_i = df_call_more['At Close'].sum()
    opt_without_money_call_j = df_call_more['Change'].sum()
    opt_without_money_put_i = df_put_less['At Close'].sum()
    opt_without_money_put_j = df_put_less['Change'].sum()

    # Итого
    fob_call_at_close_sum = df_call['At Close'].sum()
    fob_put_at_close_sum = df_put['At Close'].sum()

    SIP = [up_level, down_level, up_balance_level, down_balance_level, red_balance_level]
    SipKeys = ['up_level', 'down_level', 'up_balance_level', 'down_balance_level', 'red_balance_level']
    FOB = [opt_in_money_call_i, opt_in_money_call_j, opt_in_money_put_i, opt_in_money_put_j, opt_without_money_call_i, opt_without_money_call_j, opt_without_money_put_i, opt_without_money_put_j]
    FobKeys = ['opt_in_money_call_i', 'opt_in_money_call_j', 'opt_in_money_put_i', 'opt_in_money_put_j', 'opt_without_money_call_i', 'opt_without_money_call_j',
               'opt_without_money_put_i', 'opt_without_money_put_j']
    sip_json = {}
    fob_json = {}
    strike_json = {"puts": [], "calls": []}

    if currency == "EUR" or currency == "GBP" or currency == "AUD" or currency == "XAU" or currency == "XAG":
        for a, b, c, i, j, k in zip(call_a, call_i, call_j, put_a, put_i, put_j):
            obj_json_put = {
                "price": i / coefficient,
                "strike": j,
                "delta": k
            }
            obj_json_call = {
                "price": a / coefficient ,
                "strike": b,
                "delta": c
            }
            strike_json["puts"].append(obj_json_put)
            strike_json["calls"].append(obj_json_call)

    if currency == "CAD" or currency == "JPY":
        for a, b, c, i, j, k in zip(call_a, call_i, call_j, put_a, put_i, put_j):
            obj_json_put = {
                "price": (1 / i) * coefficient,
                "strike": j,
                "delta": k
            }
            obj_json_call = {
                "price": (1 / a)  * coefficient,
                "strike": b,
                "delta": c
            }
            strike_json["puts"].append(obj_json_put)
            strike_json["calls"].append(obj_json_call)

    sip_json = {SipKeys[i]: SIP[i] for i in range(0, len(SipKeys), 1)}
    fob_json = {FobKeys[i]: FOB[i] for i in range(0, len(FobKeys), 1)}

    name_file = "FOB" + "_" \
                + currency \
                + "_" \
                + str(datetime.date.isoformat(datetime.date.today())) \
                + ".json"
    json_file = {'strike': strike_json, 'sip': sip_json, 'fob': fob_json}
    path = os.path.join(os.environ['appdata'], 'MetaQuotes\\Terminal\\Common\\Files')
    with open(os.path.join(path, name_file), "w") as file:
        json.dump(json_file, file, cls=Int64Encoder)


def get_fut_fin():
    url_mm = "https://www.cftc.gov/files/dea/history/fut_fin_xls_" + str(datetime.datetime.now().year) + ".zip"
    file = requests.get(url_mm).content
    name = "fut_fin"
    with open(name, "wb") as f:
        f.write(file)
    if zipfile.is_zipfile(name):
        z = zipfile.ZipFile(name, 'r')
        z.extractall()

    path = os.getcwd() + f'\FinFutYY.xls'
    work_book = xlrd.open_workbook(path)
    names = work_book.sheet_names()
    fin_fut_json = []
    for i in names:
        sheet = work_book.sheet_by_name(i)
        rows = sheet.get_rows()
        count = 0
        for row in rows:
            if count > 0:
                currency = ''
                row_0 = str(row[0].value)
                if 'CANADIAN DOLLAR - CHICAGO MERCANTILE EXCHANGE' in row_0:
                    currency = 'CAD'
                elif 'SWISS FRANC - CHICAGO MERCANTILE EXCHANGE' in row_0:
                    currency = 'CHF'
                elif 'BRITISH POUND STERLING - CHICAGO MERCANTILE EXCHANGE' in row_0:
                    currency = 'GBP'
                elif 'JAPANESE YEN - CHICAGO MERCANTILE EXCHANGE' in row_0:
                    currency = 'JPY'
                elif 'EURO FX - CHICAGO MERCANTILE EXCHANGE' in row_0:
                    currency = 'EUR'
                elif 'AUSTRALIAN DOLLAR - CHICAGO MERCANTILE EXCHANGE' in row_0:
                    currency = 'AUD'
                # Date_In_Form_YYMMDD
                if currency != '':
                    # D'2020.06.06 10:49:20'
                    date_mql_format = '20' + str(int(row[1].value))
                    # print(date_mql_format)
                    date_mql_format = date_mql_format[0:4] + '.' + date_mql_format[4:6] + '.' + date_mql_format[6:]
                    # print(date_mql_format)
                    js_obj = {
                        'currency': currency,
                        'date': date_mql_format,
                        'long': int(row[8].value),
                        'short': int(row[9].value)
                    }
                    fin_fut_json.append(js_obj)
            count += 1

    name_file = "FinFut.json"

    path = os.path.join(os.environ['appdata'], 'MetaQuotes\\Terminal\\Common\\Files')
    with open(os.path.join(path, name_file), "w") as file:
        json.dump(fin_fut_json, file)

